import openpyxl
import xml.etree.cElementTree as elementTree
from xml.dom import minidom

class RequirementExporter:
    requirements = []

    def __init__(self,version,file_path):
        wb = openpyxl.load_workbook(f"{file_path}/High Level Requirements_{version}.xlsx")
        self.sheet = wb.active
        self.version = version

    def read_requirements(self):
        for i in range(2,self.sheet.max_row):
            category = self.sheet.cell(row=i, column=1).value
            if category is None:
                return
            section = self.sheet.cell(row=i, column=2).value
            item = self.sheet.cell(row=i, column=3).value
            level = "HL"
            req_type = self.sheet.cell(row=i, column=4).value
            name = self.sheet.cell(row=i, column=5).value
            content = self.sheet.cell(row=i, column=6).value
            depends_on = self.sheet.cell(row=i, column=7).value
            if depends_on is not None:
                depends_on = depends_on.split(";")
            else:
                depends_on = []

            req_id = f"{level}--{category}--{int(section)}.{int(item)}"

            requirement = {
                "req_id": str(req_id),
                "req_type": str(req_type),
                "name": str(name),
                "content": str(content),
                "depends_on": depends_on,
            }

            self.requirements.append(requirement)

    def create_xml(self,path):
        root = elementTree.Element("Requirements")
        root.attrib["version"] = str(self.version)
        root.attrib["hierarchy"] = "HLR"
        for requirement in self.requirements:
            requirement_element = elementTree.SubElement(root, "Requirement" )
            requirement_element.attrib["ID"] = requirement["req_id"]

            id_element = elementTree.SubElement(requirement_element, "Type")
            id_element.text = requirement["req_type"]

            name_element = elementTree.SubElement(requirement_element, "Name")
            name_element.text = requirement["name"]

            content_element = elementTree.SubElement(requirement_element, "Content")
            content_element.text = requirement["content"]

            relationship_element = elementTree.SubElement(requirement_element, "Relationships")
            for dependency in requirement["depends_on"]:
                dependency_element = elementTree.SubElement(relationship_element, "DependsOn")
                dependency_element.text = dependency

        rough_string = elementTree.tostring(root, 'utf-8')
        reparsed = minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="  ")
        with open(f"{path}/High Level Requirements_{self.version}.xml", "w") as f:
            f.write(pretty_xml)



